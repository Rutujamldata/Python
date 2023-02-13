# -*- coding: utf-8 -*-
"""
Created on Mon Feb 13 11:28:57 2023

@author: rutuja
"""


# import pandas lib as pd
import pandas as pd
# read by default 1st sheet of an excel file
dataframe1 = pd.read_excel('C:/Users/user/Downloads/Person.xlsx')
print(dataframe1)


#opepyxl is library to do operations on Excel data
import openpyxl
# Define variable to load the dataframe
dataframe = openpyxl.load_workbook("C:/Users/user/Downloads/Person.xlsx")
 
# Define variable to read sheet
dataframe1 = dataframe.active
 
# Iterate the loop to read the cell values
for row in range(0, dataframe1.max_row):
    for col in dataframe1.iter_cols(1, dataframe1.max_column):
        print(col[row].value)
        
        
 
        
'''Create xls file and write data in it'''
# import xlsxwriter module
import xlsxwriter
workbook = xlsxwriter.Workbook('hello.xlsx')
worksheet = workbook.add_worksheet()
worksheet.write('A1', 'Hello..')
worksheet.write('A2', 'Hello..')
worksheet.write('A3', 'Hello..')
worksheet.write('B1', 'Ajay')
worksheet.write('C1', 'How')
worksheet.write('D1', 'Are you?')
workbook.close()



'''write data using row operations'''
import xlsxwriter
workbook = xlsxwriter.Workbook('Example2.xlsx')
worksheet = workbook.add_worksheet()
row = 0
column = 0
 
content = ["ankit", "rahul", "priya", "harshita",
                    "sumit", "neeraj", "shivam"]
# iterating through content list
for item in content :
 
    # write operation perform
    worksheet.write(row, column, item)
 
    # incrementing the value of row by one
    # with each iterations.
    row += 1
     
workbook.close()

'''write data using row & column operations'''
import xlsxwriter
workbook = xlsxwriter.Workbook('Example3.xlsx')
worksheet = workbook.add_worksheet("My sheet")

scores = (
    ['ankit', 1000],
    ['rahul',   100],
    ['priya',  300],
    ['harshita',    50],
)
 
# Start from the first cell. Rows and
# columns are zero indexed.
row = 0
col = 0
 
# Iterate over the data and write it out row by row.
for name, score in (scores):
    worksheet.write(row, col, name)
    worksheet.write(row, col + 1, score)
    row += 1
 
workbook.close()
